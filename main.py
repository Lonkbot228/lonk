import os
import re
import logging
import datetime
from io import BytesIO

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from telegram import Update
from telegram.constants import ParseMode
from telegram.ext import Application, CommandHandler, ContextTypes

# Настройки логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

# URL публичной папки Google Диска (замените на вашу ссылку)
PUBLIC_FOLDER_URL = "https://drive.google.com/drive/folders/1kUYiSAafghhYR0ARyXwPW1HZPpHcFIag""

# Регулярное выражение для имени файла вида DD.MM.YYYY.xlsx
FILENAME_REGEX = re.compile(r'(\d{2}\.\d{2}\.\d{4}\.xlsx)')

def get_candidate_dates():
    """Возвращает список дат в нужном порядке для поиска файла."""
    today = datetime.date.today()
    # 1. Сегодня
    dates = [today]
    # 2. Завтра и послезавтра
    tomorrow = today + datetime.timedelta(days=1)
    day_after = today + datetime.timedelta(days=2)
    dates.extend([tomorrow, day_after])
    # 3. Вчера
    yesterday = today - datetime.timedelta(days=1)
    dates.append(yesterday)
    return dates

def format_date_for_filename(date_obj: datetime.date) -> str:
    """Форматирует дату в строку DD.MM.YYYY.xlsx"""
    return date_obj.strftime("%d.%m.%Y") + ".xlsx"

def fetch_folder_page():
    """Получает HTML страницы публичной папки Google Диска."""
    try:
        response = requests.get(PUBLIC_FOLDER_URL)
        response.raise_for_status()
        return response.text
    except Exception as e:
        logger.error("Ошибка получения страницы папки: %s", e)
        return None

def parse_files_from_folder(html_text):
    """
    Из HTML-страницы получает словарь: {имя_файла: (file_id, download_url)}.
    Предполагается, что ссылки на файлы содержат шаблон /file/d/FILE_ID/view
    и где-то рядом присутствует имя файла.
    """
    files = {}
    soup = BeautifulSoup(html_text, "lxml")
    # Ищем все ссылки, в которых встречается /file/d/
    for a in soup.find_all("a", href=True):
        href = a["href"]
        match = re.search(r'/file/d/([a-zA-Z0-9_-]+)/', href)
        if match:
            file_id = match.group(1)
            text = a.get_text(strip=True)
            # Ищем имя файла по регулярному выражению
            fname_match = FILENAME_REGEX.search(text)
            if fname_match:
                file_name = fname_match.group(1)
                # Формируем ссылку для скачивания
                download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
                files[file_name] = (file_id, download_url)
    return files

def find_schedule_file():
    """
    Поиск файла с расписанием по алгоритму:
      1. Ищем файл за сегодня.
      2. Если нет, ищем за завтрашний или послезавтрашний день.
      3. Если нет, ищем за вчерашний день.
    Возвращает кортеж (имя_файла, download_url) или (None, None).
    """
    html_text = fetch_folder_page()
    if not html_text:
        return None, None

    files = parse_files_from_folder(html_text)
    candidate_dates = get_candidate_dates()
    for date_obj in candidate_dates:
        fname = format_date_for_filename(date_obj)
        if fname in files:
            _, download_url = files[fname]
            logger.info("Найден файл: %s", fname)
            return fname, download_url
    logger.info("Файл расписания не найден по заданным датам.")
    return None, None

def download_file(download_url):
    """Скачивает файл по ссылке и возвращает его содержимое в виде BytesIO."""
    try:
        response = requests.get(download_url)
        response.raise_for_status()
        return BytesIO(response.content)
    except Exception as e:
        logger.error("Ошибка скачивания файла: %s", e)
        return None

def extract_schedule_from_workbook(wb):
    """
    Обрабатывает книгу .xlsx.
    Для каждого листа, название которого начинается с "Пара",
    ищется строка, содержащая "СА-17". Если найдено:
      - Если во 2-м столбце содержится "СА-17", берем данные из 1-го и 3-го столбцов.
      - Если в 5-м столбце содержится "СА-17", берем данные из 4-го и 6-го столбцов.
    Возвращает словарь вида:
       { sheet_title: {"room": значение, "teacher": значение} }
    """
    schedule = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Пара"):
            continue
        ws = wb[sheet_name]
        room = None
        teacher = None
        for row in ws.iter_rows(values_only=True):
            if row is None or len(row) < 6:
                continue
            # Проверяем наличие "СА-17" в 2-м столбце (индекс 1)
            if row[1] is not None and "СА-17" in str(row[1]):
                part1 = str(row[0]).strip() if row[0] is not None else ""
                part2 = str(row[2]).strip() if row[2] is not None else ""
                room = (part1 + part2).strip()
            # Проверяем наличие "СА-17" в 5-м столбце (индекс 4)
            if row[4] is not None and "СА-17" in str(row[4]):
                part1 = str(row[3]).strip() if row[3] is not None else ""
                part2 = str(row[5]).strip() if row[5] is not None else ""
                teacher = " ".join(filter(None, [part1, part2])).strip()
            # Если нашли хотя бы одно из значений, можно завершить поиск в этом листе
            if room or teacher:
                break
        if room or teacher:
            schedule[sheet_name] = {"room": room, "teacher": teacher}
    return schedule

def format_schedule_message(schedule_data, file_date: datetime.date):
    """
    Форматирует итоговое сообщение согласно шаблону:
    
    🗓️ **2 апреля**  

    📎 **Пара 1**  
    🔑 103  
    ✍️ Старых О.А.  
    """
    # Форматируем дату, например, "2 апреля"
    day = file_date.day
    month = file_date.strftime("%B")  # название месяца на английском; можно заменить на русские названия
    # Для примера заменим английские названия на русские (можно расширить)
    months_ru = {
        "January": "января", "February": "февраля", "March": "марта",
        "April": "апреля", "May": "мая", "June": "июня",
        "July": "июля", "August": "августа", "September": "сентября",
        "October": "октября", "November": "ноября", "December": "декабря"
    }
    month_ru = months_ru.get(month, month)
    message_lines = [f"🗓️ **{day} {month_ru}**", ""]
    # Для каждого листа ("Пара X")
    for sheet, data in schedule_data.items():
        message_lines.append(f"📎 **{sheet}**")
        if data.get("room"):
            message_lines.append(f"🔑 {data['room']}")
        if data.get("teacher"):
            message_lines.append(f"✍️ {data['teacher']}")
        message_lines.append("")  # пустая строка между парами
    return "\n".join(message_lines)

async def schedule_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /schedule."""
    # Ищем файл расписания
    fname, download_url = find_schedule_file()
    if not fname or not download_url:
        await update.message.reply_text("Файл расписания не найден.")
        return

    # Определяем дату файла из имени (формат DD.MM.YYYY.xlsx)
    try:
        file_date = datetime.datetime.strptime(fname[:-5], "%d.%m.%Y").date()
    except Exception as e:
        file_date = datetime.date.today()

    file_stream = download_file(download_url)
    if not file_stream:
        await update.message.reply_text("Ошибка скачивания файла.")
        return

    try:
        wb = load_workbook(filename=file_stream, data_only=True)
    except Exception as e:
        logger.error("Ошибка открытия файла: %s", e)
        await update.message.reply_text("Ошибка обработки файла.")
        return

    schedule_data = extract_schedule_from_workbook(wb)
    if not schedule_data:
        await update.message.reply_text("Не найдено расписание для СА-17.")
        return

    message = format_schedule_message(schedule_data, file_date)
    await update.message.reply_text(message, parse_mode=ParseMode.MARKDOWN)

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start."""
    welcome_text = "Привет! Используй команду /schedule для получения расписания."
    await update.message.reply_text(welcome_text)

def main():
    # Получаем токен из переменной окружения
    token = os.environ.get("TOKEN")
    if not token:
        logger.error("Переменная окружения TOKEN не установлена")
        return

    # Создаем приложение бота с использованием токена из переменной окружения
    application = Application.builder().token(token).build()

    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("schedule", schedule_command))

    application.run_polling()

if __name__ == '__main__':
    main()
